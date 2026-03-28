# mars/analysis/report.py

import os
import sys
from copy import copy
from importlib import resources
import polars as pl
import pandas as pd
import numpy as np
from typing import Dict, Tuple, Optional, Union, List, Any, NamedTuple
from mars.utils.logger import logger

try:
    from IPython.display import display, HTML
except ImportError:
    display = None
    
class ProfileData(NamedTuple):
    overview: Union[pl.DataFrame, pd.DataFrame]
    dq_trends: Dict[str, Union[pl.DataFrame, pd.DataFrame]]
    stats_trends: Dict[str, Union[pl.DataFrame, pd.DataFrame]]

class MarsProfileReport:
    """
    由 MarsDataProfiler 生成的数据画像报告容器。

    该类作为数据探查 (EDA) 流程的最终输出枢纽，负责统一管理特征统计指标与数据质量 (DQ) 指标的
    展示、交互与导出。它无缝连接了底层纯粹的分析数据与业务级可视化报表。

    核心特性
    --------
    * **交互式探查 (Interactive EDA)**: 在 Jupyter Notebook 中渲染富文本 HTML，
      支持嵌入式迷你分布图 (Sparklines)、热力图色阶与动态数据表。
    * **企业级导出 (Professional Export)**: 一键生成符合业务阅读直觉的 Excel 监控报表，
      并原生保留所有条件格式 (Conditional Formatting)、数据条与百分比刻度。
    * **趋势追踪 (Trend Tracking)**: 管理按时间切片或客群维度展开的多维分析结果，
      通过统一的 API 快速下钻追踪特定指标的演变趋势。

    Attributes
    ----------
    overview_table : Union[pl.DataFrame, pd.DataFrame]
        全量特征概览宽表。包含所有特征的全局数据质量与统计分布表现 (如 missing_rate, mean 等)。
    dq_tables : Dict[str, Union[pl.DataFrame, pd.DataFrame]]
        数据质量 (DQ) 指标的分组趋势字典。
        - Key: 指标名称 (如 'missing', 'zeros', 'unique')。
        - Value: 该指标在不同时间切片/分组下的透视宽表 (Pivot Table)。
    stats_tables : Dict[str, Union[pl.DataFrame, pd.DataFrame]]
        统计分布指标的分组趋势字典。
        - Key: 指标名称 (如 'mean', 'max', 'p25')。
        - Value: 该指标在不同时间切片/分组下的透视宽表 (Pivot Table)。

    Examples
    --------
    >>> from mars.analysis import MarsDataProfiler
    >>> profiler = MarsDataProfiler(df)
    >>> report = profiler.generate_profile(profile_by="month")
    >>> 
    >>> # 1. 在 Jupyter 中进行交互式展示
    >>> report.show_overview(sort_by="missing_rate")
    >>> report.show_trend("missing", features=["age", "income"])
    >>> 
    >>> # 2. 提取底层干净的分析数据进行二次开发
    >>> overview_df, dq_dict, stat_dict = report.get_profile_data()
    >>> 
    >>> # 3. 导出为带高亮业务样式的 Excel 报表
    >>> report.write_excel("mars_data_health_audit.xlsx")
    """

    def __init__(
        self, 
        overview: Union[pl.DataFrame, pd.DataFrame],
        dq_tables: Dict[str, Union[pl.DataFrame, pd.DataFrame]],
        stats_tables: Dict[str, Union[pl.DataFrame, pd.DataFrame]]
    ) -> None:
        self.overview_table = overview
        self.dq_tables = dq_tables
        self.stats_tables = stats_tables
        
        # 建立索引：将所有指标名映射到对应的数据源类型 ('dq' 或 'stat')
        # 这允许我们在 show_trend 中快速定位
        self._metric_index: Dict[str, str] = {}
        for k in self.dq_tables.keys():
            self._metric_index[k] = "dq"
        for k in self.stats_tables.keys():
            self._metric_index[k] = "stat"

    def get_profile_data(self) -> ProfileData:
        """[API] 获取原始数据对象"""
        return ProfileData(
            overview=self.overview_table,
            dq_trends=self.dq_tables,
            stats_trends=self.stats_tables
        )

    def _repr_html_(self) -> str:
        """
        [Internal] Jupyter Notebook 控制面板 
        """
        df_ov = self.overview_table
        n_feats = len(df_ov) if hasattr(df_ov, "__len__") else df_ov.height
        
        dq_keys = list(self.dq_tables.keys())
        stat_keys = list(self.stats_tables.keys())

        # 样式定义 (Inline CSS for portability)
        # 胶囊样式，用于包裹指标名
        pill_style = (
            "background-color: #e8f4f8; color: #2980b9; border: 1px solid #bce0eb; "
            "padding: 2px 6px; border-radius: 4px; font-family: monospace; font-size: 0.9em; margin-right: 4px;"
        )
        # 代码块样式
        code_style = (
            "background-color: #f0f0f0; padding: 2px 4px; border-radius: 3px; "
            "font-family: monospace; color: #e74c3c; font-weight: bold;"
        )
        
        # 辅助函数：生成指标徽章列表
        def _fmt_pills(keys):
            if not keys: return "<span style='color:#ccc'>None</span>"
            # 为了防止指标太多撑爆屏幕，限制显示数量 (例如只显示前 20 个，后面加 ...)
            display_keys = keys[:30] 
            pills = "".join([f"<span style='{pill_style}'>'{k}'</span>" for k in display_keys])
            if len(keys) > 30:
                pills += f"<span style='color:#999; font-size:0.8em'> (+{len(keys)-30} more)</span>"
            return pills

        # 组装 HTML
        return f"""
        <div style="border: 1px solid #e0e0e0; border-left: 5px solid #2980b9; border-radius: 4px; background: white; max-width: 900px; font-family: 'Segoe UI', sans-serif;">
            
            <div style="padding: 12px 15px; background-color: #f8f9fa; border-bottom: 1px solid #e0e0e0; display: flex; justify-content: space-between; align-items: center;">
                <div style="font-weight: bold; color: #2c3e50; font-size: 1.1em;">
                    📊 Mars Data Profile
                </div>
                <div style="font-size: 0.85em; color: #7f8c8d;">
                    <span style="margin-left: 15px;">🏷️ Features: <b>{n_feats}</b></span>
                    <span style="margin-left: 15px;">🔍 DQ Metrics: <b>{len(dq_keys)}</b></span>
                    <span style="margin-left: 15px;">📉 Stat Metrics: <b>{len(stat_keys)}</b></span>
                </div>
            </div>

            <div style="padding: 15px;">
                
                <div style="margin-bottom: 15px;">
                    <div style="font-size: 0.8em; text-transform: uppercase; color: #95a5a6; font-weight: bold; margin-bottom: 5px;">Quick Actions</div>
                    <div style="display: flex; gap: 20px; font-size: 0.95em;">
                        <div>👉 <span style="{code_style}">.show_overview()</span> &nbsp;<span style="color:#555">View Full Report</span></div>
                        <div>💾 <span style="{code_style}">.write_excel()</span> &nbsp;<span style="color:#555">Export XLSX</span></div>
                        <div>📥 <span style="{code_style}">.get_profile_data()</span> &nbsp;<span style="color:#555">Get Raw Data</span></div>
                    </div>
                </div>

                <div style="border-top: 1px dashed #e0e0e0; padding-top: 12px;">
                    <div style="font-size: 0.8em; text-transform: uppercase; color: #95a5a6; font-weight: bold; margin-bottom: 8px;">
                        Trend Analysis <span style="font-weight:normal; text-transform:none; color:#bbb">(Use <code>.show_trend('metric_name')</code>)</span>
                    </div>
                    
                    <div style="display: flex; margin-bottom: 8px; align-items: baseline;">
                        <div style="width: 80px; font-weight: bold; color: #27ae60; font-size: 0.9em;">DQ:</div>
                        <div style="flex: 1; line-height: 1.6;">
                            {_fmt_pills(dq_keys)}
                        </div>
                    </div>
                    
                    <div style="display: flex; align-items: baseline;">
                        <div style="width: 80px; font-weight: bold; color: #2980b9; font-size: 0.9em;">Stats:</div>
                        <div style="flex: 1; line-height: 1.6;">
                            {_fmt_pills(stat_keys)}
                        </div>
                    </div>
                </div>

            </div>
            
            <div style="padding: 6px 15px; background-color: #fff8e1; border-top: 1px solid #fae5b0; font-size: 0.8em; color: #d35400;">
                💡 <b>Pro Tip:</b> Use <span style="{code_style}">.show_trend('psi')</span> to detect population stability drift.
            </div>
        </div>
        """

    def show_overview(self, 
                      features: Optional[Union[str, List[str]]] = None, 
                      sort_by: Optional[Union[str, List[str]]] = None, 
                      sort_ascending: bool = False) -> "pd.io.formats.style.Styler":
        """
        展示全量概览大宽表。

        Parameters
       -------
        features : str or List[str], optional
            需要展示的特征名列表。若为 None，则展示所有特征。
        sort_by : str or List[str], optional
            排序的依据列。默认先按 dtype 聚类，再按 missing_rate 排序。
        sort_ascending : bool, default False
            排序方向。默认降序 (False)，即把问题最严重的特征排在前面。
        """
        # 转换为 Pandas 副本以进行切片
        df = self._to_pd(self.overview_table).copy()
        
        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]

        return self._get_styler(
            df,
            title="Dataset Overview", 
            cmap="RdYlGn_r", 
            sort_by= ["dtype"] + (["missing_rate"] if sort_by is None else ([sort_by] if isinstance(sort_by, str) else sort_by)),
            sort_ascending=sort_ascending, 
            # 指定哪些列应用“红绿灯”配色 (高值=红)
            subset_cols=["missing_rate", "zeros_rate", "unique_rate", "top1_ratio"],
            fmt_as_pct=False # 概览表混合了多种类型，不强制全转百分比，由内部逻辑细分
        )

    def show_trend(self, 
                   metric: str, 
                   features: Optional[Union[str, List[str]]] = None, 
                   group_ascending: bool = True, 
                   sort_by: Union[List[str], str] = "total", 
                   sort_ascending: bool = False) -> "pd.io.formats.style.Styler":
        """
        展示指定指标的分组趋势。

        Parameters
       -------
        metric : str
            指标名称 (如 'missing', 'mean', 'psi')。
        features : str or List[str], optional
            需要展示的特征名列表。若为 None，则展示所有特征。
        group_ascending : bool, default True
            分组/时间切片列的排序方向 (横向)。True 表示正序（从左到右由旧到新）。
        sort_by : str or List[str], default 'total'
            趋势表内部排序的依据列，可以是单列名或多列名列表。
        sort_ascending : bool, default False
            趋势表内部排序的方式，默认为降序。
        """
        # 路由逻辑：查找指标属于哪个表
        source_type = self._metric_index.get(metric)
        if source_type is None:
             # 提供更友好的报错提示
            available = list(self._metric_index.keys())
            raise ValueError(f"❌ Metric '{metric}' not found. Available metrics: {available[:10]}...")

        # 获取数据
        if source_type == "dq":
            df_raw = self.dq_tables[metric]
            # DQ 默认配置
            cmap = "RdYlGn_r"  # 红色代表高风险 (高缺失)
            fmt_pct = True     # DQ 指标通常是率 (Rate/Ratio)
            vmin, vmax = 0, 1  # 率通常在 0~1 之间
            
        else: # source_type == "stat"
            df_raw = self.stats_tables[metric]
            # Stats 默认配置
            cmap = "Blues"     # 蓝色代表数值高低 (中性)
            fmt_pct = False    # 统计值通常是绝对值
            vmin, vmax = None, None

        # 特殊指标微调 (Override)
        if metric == "psi":
            cmap = "RdYlGn_r" # PSI 高了是坏事
            fmt_pct = False   # PSI 是数值不是百分比
            vmin, vmax = 0.0, 0.5 # 锚定阈值
        
        df = self._to_pd(df_raw).copy()

        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]

        # 排序
        df = df.sort_values(by=sort_by, ascending=sort_ascending)
        df = self._reorder_trend_cols(df, group_ascending=group_ascending)

        return self._get_styler(
            df,
            title=f"Trend Analysis: {metric}",
            cmap=cmap,
            fmt_as_pct=fmt_pct,
            vmin=vmin, 
            vmax=vmax,
            add_bars=True # 所有趋势表都允许显示 CV 条
        )

    def _reorder_trend_cols(self, df: pd.DataFrame, group_ascending: bool) -> pd.DataFrame:
        """[Internal Helper] 重新排列趋势表的列顺序。"""
        # 定义元数据列和末尾统计列
        meta_cols = ["feature", "dtype", "distribution", "top1_value"]
        stat_cols = ["total", "group_mean", "group_var", "group_cv"]
        
        # 识别中间的分组列（如时间列）
        all_cols = df.columns.tolist()
        group_cols = [c for c in all_cols if c not in meta_cols + stat_cols]
        
        # 排序分组列 (受 group_ascending 控制)
        group_cols_sorted = sorted(group_cols, reverse=not group_ascending)
        
        # 组合最终顺序
        final_order = [c for c in meta_cols if c in all_cols] + \
                      group_cols_sorted + \
                      [c for c in stat_cols if c in all_cols]
        return df[final_order]

    def write_excel(self, 
                    path: str = "mars_report.xlsx", 
                    group_ascending: bool = True,
                    sort_by: Union[str, List[str]] = "total",
                    sort_ascending: bool = False) -> None:
        """
        导出带有精美样式和趋势热力图的 Excel 报告。
        """
        logger.info(f"📊 Exporting report to: {path}...")
        
        # 1. 依赖检查
        try:
            import xlsxwriter
        except ImportError:
            logger.error("❌ 'xlsxwriter' is required for Excel export. Install it via: pip install xlsxwriter")
            return

        try:
            with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
                #--------------------------------------------------------
                # 1. 导出概览页 (Overview)
                #--------------------------------------------------------
                overview_styler = self.show_overview()
                if overview_styler is not None:
                    overview_styler.to_excel(writer, sheet_name="Overview", index=False)
                
                #--------------------------------------------------------
                # 2. 统一导出所有趋势页 (Trend & DQ)
                #--------------------------------------------------------
                dq_keys = list(self.dq_tables.keys())
                stat_keys = list(self.stats_tables.keys())
                all_metrics = dq_keys + stat_keys
                
                for metric in all_metrics:
                    # 保证导出的表结构与 Notebook 展示完全一致
                    styler = self.show_trend(
                        metric, 
                        group_ascending=group_ascending, 
                        sort_by=sort_by, 
                        sort_ascending=sort_ascending
                    )
                    
                    if styler is not None:
                        prefix = "DQ" if metric in self.dq_tables else "Trend"
                        sheet_name = f"{prefix}_{metric.capitalize()}"[:31]
                        
                        styler.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # 确保条件格式锚定的列与导出的表完全吻合
                        self._apply_excel_formatting(
                            writer, sheet_name, metric, 
                            group_ascending=group_ascending,
                            sort_by=sort_by,
                            sort_ascending=sort_ascending
                        )

                # 3. 自动列宽调整
                for sheet in writer.sheets.values():
                    sheet.autofit()
                    
            logger.info("Report exported successfully.")

        except Exception as e:
            logger.error(f"❌ Failed to export Excel: {e}", exc_info=True)

    def _apply_excel_formatting(self, 
                                writer, 
                                sheet_name: str, 
                                metric: str, 
                                group_ascending: bool,
                                sort_by: Union[str, List[str]],
                                sort_ascending: bool):
        """
        [Helper] 抽离 Excel 条件格式逻辑，保持主流程清晰。
        """
        if metric in self.dq_tables:
            raw_df = self.dq_tables[metric]
        else:
            raw_df = self.stats_tables[metric]
            
        # 必须和 show_trend 的内部重排逻辑一模一样，否则 Excel 样式会错位
        df_pd: pd.DataFrame = self._to_pd(raw_df).copy()
        
        # 匹配行排序
        if sort_by in df_pd.columns or (isinstance(sort_by, list) and all(c in df_pd.columns for c in sort_by)):
            df_pd = df_pd.sort_values(by=sort_by, ascending=sort_ascending)
            
        # 匹配列排序
        df_pd = self._reorder_trend_cols(df_pd, group_ascending=group_ascending)
        
        # [优化] 动态查找时间列的索引范围
        meta_and_stat = set(["feature", "dtype", "distribution", "top1_value", "total", "group_mean", "group_var", "group_cv"])
        time_cols = [c for c in df_pd.columns if c not in meta_and_stat]
        
        if not time_cols:
            return

        worksheet = writer.sheets[sheet_name]
        
        # PSI 专用三色阶 (红绿灯)
        if metric == "psi":
            meta_cols = ["feature", "dtype", "distribution", "top1_value"]
            start_col = 0
            for i, col in enumerate(df_pd.columns):
                if col not in meta_cols:
                    start_col = i
                    break
            
            end_col = len(df_pd.columns) - 1
            
            worksheet.conditional_format(1, start_col, len(df_pd), end_col, {
                'type': '3_color_scale',
                'min_type': 'num', 'min_value': 0.05, 'min_color': '#63BE7B', # Green
                'mid_type': 'num', 'mid_value': 0.15, 'mid_color': '#FFEB84', # Yellow
                'max_type': 'num', 'max_value': 0.25, 'max_color': '#F8696B'  # Red
            })

        # 稳定性 Data Bars (针对 group_cv)
        if "group_cv" in df_pd.columns:
            col_idx = df_pd.columns.get_loc("group_cv")
            worksheet.conditional_format(1, col_idx, len(df_pd), col_idx, {
                'type': 'data_bar', 
                'bar_color': '#638EC6', 
                'bar_solid': True,
                'min_type': 'num', 'min_value': 0, 
                'max_type': 'num', 'max_value': 1
            })

    def _to_pd(self, df: Any) -> pd.DataFrame:
        """
        [辅助方法] 确保数据转换为 Pandas DataFrame 格式。
        """
        if isinstance(df, pl.DataFrame):
            return df.to_pandas()
        return df

    def _get_styler(
        self, 
        df_input: Any, 
        title: str, 
        cmap: str, 
        sort_by: Optional[List[str]] = None,
        sort_ascending: bool = False, # 统一内部 API 命名
        subset_cols: Optional[List[str]] = None, 
        add_bars: bool = False, 
        fmt_as_pct: bool = False,
        vmin: Optional[float] = None, 
        vmax: Optional[float] = None
    ) -> Optional["pd.io.formats.style.Styler"]:
        """
        [Internal] 通用样式生成器。
        """
        if df_input is None:
            return None
        df: pd.DataFrame = self._to_pd(df_input)
        if sort_by is not None:
            df = df.sort_values(by=sort_by, ascending=sort_ascending) # 使用统一参数进行底层排序
        if df.empty:
            return None

        # 元数据排除列表
        exclude_meta: List[str] = [
            "feature", "dtype", 
            "group_mean", "group_var", "group_cv",
            "distribution",
            "top1_value"
            ]
        
        # 确定色彩渐变范围
        if subset_cols:
            gradient_cols: List[str] = [c for c in subset_cols if c in df.columns]
        else:
            gradient_cols = [c for c in df.columns if c not in exclude_meta]

        styler = df.style.set_caption(f"<b>{title}</b>").hide(axis="index")
        
        # 应用热力图
        if gradient_cols:
            styler = styler.background_gradient(
                cmap=cmap, 
                subset=gradient_cols, 
                axis=None,
                vmin=vmin,
                vmax=vmax
            )
        
        # 应用数据条
        if add_bars and "group_cv" in df.columns:
            styler = styler.bar(subset=["group_cv"], color='#ff9999', vmin=0, vmax=1, width=90)
            styler = styler.format("{:.4f}", subset=["group_cv", "group_var"])

        # 数值格式化逻辑
        num_cols: pd.Index = df.select_dtypes(include=['number']).columns
        data_cols: List[str] = [c for c in num_cols if c not in ["group_var", "group_cv", "distribution"]]

        pct_format: str = "{:.2%}"  
        float_format: str = "{:.2f}"

        if fmt_as_pct:
            if data_cols:
                styler = styler.format(pct_format, subset=data_cols)
        else:
            pct_cols: List[str] = [
                c for c in df.columns 
                if ("rate" in c or "ratio" in c) and (c in num_cols)
            ]
            
            if pct_cols:
                styler = styler.format(pct_format, subset=pct_cols)
            
            float_cols: List[str] = [c for c in data_cols if c not in pct_cols]
            if float_cols:
                styler = styler.format(float_format, subset=float_cols)
        
        # 分布迷你图样式
        if "distribution" in df.columns:
            styler = styler.set_table_styles([
                {'selector': '.col_distribution', 'props': [
                    # 优先使用 Consolas (Win) 或 Menlo (Mac)，最后 fallback 到 monospace
                    ('font-family', '"Consolas", "Menlo", "Courier New", monospace'), 
                    ('color', '#1f77b4'),
                    ('white-space', 'pre'), # [关键] 防止 HTML 自动压缩连续空格
                    ('font-weight', 'bold'),
                    ('text-align', 'left')
                ]}
            ], overwrite=False)

        # 全局表格外观
        styler = styler.set_table_styles([
            {
                'selector': 'th', 
                'props': [('text-align', 'left'), ('background-color', '#f0f2f5'), ('color', '#333')]
            },
            {
                'selector': 'caption', 
                'props': [('font-size', '1.2em'), ('padding', '10px 0'), ('color', '#2c3e50')]
            }
        ], overwrite=False)

        return styler
    
class MarsEvaluationReport:
    """
    由 MarsBinEvaluator 生成的特征效能与稳定性评估报告容器。

    该类是风控特征工程的核心交付物载体。它不仅负责安全存储底层极速计算出的海量评估数据，
    还提供了一套为金融级建模场景深度定制的分析与交互接口。它智能兼容 Polars 与 Pandas 
    双引擎，确保数据流水线的类型连贯性。

    核心特性
    --------
    * **交互式审计 (Interactive Audit)**: 在 Jupyter 环境中渲染带业务语境色彩 
      (如 RdYlGn_r 预警色带) 的富文本 Styler，快速扫描特征区分度与单调性缺陷。
    * **时序趋势追踪 (Time-Series Tracking)**: 动态聚合生成 PSI、AUC、IV、坏账率及逻辑稳定性
      (RiskCorr) 的跨期趋势热力图，精准定位特征分布漂移 (Data Drift) 的时间拐点。
    * **生产级报表导出 (Production-Ready Export)**: 一键生成包含条件格式 (红绿灯/数据条)
      和专业分箱排版的跨平台多 Sheet Excel 监控月报，实现从代码到业务汇报的无缝衔接。

    Attributes
    ----------
    summary_table : Union[pl.DataFrame, pd.DataFrame]
        特征级汇总审计宽表。涵盖全局预测力 (如 Total IV, Max KS, AUC) 与
        跨期稳定性边界 (如 Max PSI, Min RiskCorr) 的核心雷达数据。
    trend_tables : Dict[str, Union[pl.DataFrame, pd.DataFrame]]
        核心评估指标的跨期趋势字典。
        - Key: 指标名称 (如 'psi', 'auc', 'iv', 'bad_rate', 'risk_corr')。
        - Value: 结构为 [特征名 x 时间切片] 的透视宽表 (Pivot Table)。
    detail_table : Union[pl.DataFrame, pd.DataFrame]
        最细粒度的分箱明细表。包含每个特征、每个时间切片下所有分箱的样本分布占比、
        坏账率、Lift、WOE 及累积风险推演指标。
    group_col : str, optional
        驱动趋势分析的切片维度标识 (如 'month', 'vintage', 'channel')。
        若处于单点快照评估模式 (Snapshot Mode)，则通常为 'Total' 或 None。

    Examples
    --------
    >>> from mars.analysis import MarsBinEvaluator
    >>> evaluator = MarsBinEvaluator(target="is_bad")
    >>> report = evaluator.evaluate(df, profile_by="month")
    >>> 
    >>> # 1. 快速查看核心特征的全局排行与红绿灯预警
    >>> core_features = ["age", "debt_ratio", "revolving_util"]
    >>> report.show_summary(features=core_features)
    >>> 
    >>> # 2. 下钻追踪特定指标的时间序列漂移轨迹 (按全局表现降序，时间正序)
    >>> report.show_trend("psi", sort_by="Total", sort_ascending=False, group_ascending=True)
    >>> 
    >>> # 3. 导出包含全量明细和色彩高亮的专业 Excel 监控报表
    >>> report.write_excel("mars_feature_evaluation.xlsx")
    """

    def __init__(
        self, 
        summary_table: Union[pl.DataFrame, pd.DataFrame],
        trend_tables: Dict[str, Union[pl.DataFrame, pd.DataFrame]],
        detail_table: Union[pl.DataFrame, pd.DataFrame],
        group_col: Optional[str] = None
    ) -> None:
        """
        初始化报告容器。

        Parameters
       -------
        summary_table : Union[pl.DataFrame, pd.DataFrame]
            特征级汇总表。
        trend_tables : Dict[str, Union[pl.DataFrame, pd.DataFrame]]
            指标趋势表字典。
        detail_table : Union[pl.DataFrame, pd.DataFrame]
            最细粒度的分箱明细表。
        group_col : str, optional
            分组列名（例如：'month' 或 'vintage'）。
        """
        # 直接存储原始数据，不再强制命名为 _pl，以支持多种类型
        self._summary = summary_table
        self._trend_dict = trend_tables
        self._detail = detail_table
        self.group_col = group_col 
        
    @property
    def summary_table(self) -> Union[pl.DataFrame, pd.DataFrame]:
        """返回汇总统计表（类型与输入一致）。"""
        return self._summary

    @property
    def trend_tables(self) -> Dict[str, Union[pl.DataFrame, pd.DataFrame]]:
        """返回趋势宽表字典（类型与输入一致）。"""
        return self._trend_dict

    @property
    def detail_table(self) -> Union[pl.DataFrame, pd.DataFrame]:
        """返回分箱明细表（类型与输入一致）。"""
        return self._detail

    def get_evaluation_data(self) -> Tuple[
        Union[pl.DataFrame, pd.DataFrame], 
        Dict[str, Union[pl.DataFrame, pd.DataFrame]], 
        Union[pl.DataFrame, pd.DataFrame]
    ]:
        """
        获取所有原始数据。
        
        Returns
       ----
        Tuple
            返回 (汇总表, 趋势表字典, 明细表)，类型与输入一致。
        """
        return self.summary_table, self.trend_tables, self.detail_table

    def _to_pd(self, df: Any) -> pd.DataFrame:
        """辅助函数：将输入对象转为 Pandas DataFrame（用于展示或导出）。"""
        if isinstance(df, pl.DataFrame):
            return df.to_pandas()
        return df

    def _repr_html_(self) -> str:
        """
        [Dashboard] Jupyter Notebook 控制面板。
        """
        # 内部展示逻辑统一转为 Pandas 处理
        df_summary_pd = self._to_pd(self.summary_table)
        n_feats = len(df_summary_pd)
        
        # 简单统计报警数 (修正为新的小写列名 psi_max)
        high_risk_psi = 0
        if "psi_max" in df_summary_pd.columns:
            high_risk_psi = sum(df_summary_pd["psi_max"] > 0.25)

        # 样式定义
        pill_style = (
            "background-color: #e8f4f8; color: #2980b9; border: 1px solid #bce0eb; "
            "padding: 2px 6px; border-radius: 4px; font-family: monospace; font-size: 0.9em; margin-right: 4px;"
        )
        
        # 动态生成 Trend 链接
        trend_keys = list(self.trend_tables.keys())
        trend_pills = "".join([f"<span style='{pill_style}'>'{k}'</span>" for k in trend_keys])

        lines = []
        # 查看类操作
        lines.append('👉 <code>.show_summary()</code> &nbsp;<span style="color:#7f8c8d">View Feature Ranking</span>')
        lines.append(f'👉 <code>.show_trend(metric)</code> <span style="color:#7f8c8d">metric: {trend_pills}</span>')
        
        # [新增] 获取数据类操作
        lines.append('<hr style="margin: 8px 0; border: 0; border-top: 1px dashed #ccc;">')
        lines.append('📥 <code>.get_evaluation_data()</code> &nbsp;<span style="color:#7f8c8d">Get Raw Data (summary, trends, detail)</span>')
        lines.append('💾 <code>.write_excel()</code> &nbsp;<span style="color:#7f8c8d">Export to Excel</span>')

        return f"""
        <div style="border-left: 5px solid #8e44ad; background-color: #f4f6f7; padding: 15px; border-radius: 0 5px 5px 0; font-family: 'Segoe UI', sans-serif;">
            <h3 style="margin:0 0 10px 0; color:#2c3e50;">📉 Mars Feature Evaluation</h3>
            
            <div style="display: flex; gap: 30px; margin-bottom: 12px; font-size: 0.95em;">
                <div><strong>🏷️ Features:</strong> {n_feats}</div>
                <div><strong>🚨 High PSI (>0.25):</strong> <span style="color: {'red' if high_risk_psi > 0 else 'green'}; font-weight:bold;">{high_risk_psi}</span></div>
                <div><strong>📅 Group By:</strong> {self.group_col if self.group_col else 'None (Total Only)'}</div>
            </div>
            
            <div style="font-size:0.9em; line-height:1.8; color:#2c3e50; background: white; padding: 10px; border: 1px solid #e0e0e0; border-radius: 4px;">
                { "<br>".join(lines) }
            </div>
        </div>
        """

    def show_summary(self, 
                     features: Optional[Union[str, List[str]]] = None # 新增特征筛选
                     ) -> "pd.io.formats.style.Styler":
        """
        展示特征汇总评分表。

        Parameters
       -------
        features : str or List[str], optional
            需要展示的特征名列表。若为 None，则展示所有特征。
        """
        df: pd.DataFrame = self._to_pd(self.summary_table).copy()
        
        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]
        
        # [UI 优化] 如果是多目标模式，自动将 target 列提取到最前面
        for t_col in ["target", "target_col", "y"]:
            if t_col in df.columns:
                cols = [t_col] + [c for c in df.columns if c != t_col]
                df = df[cols]
                break

        styler = df.style.set_caption("<b>Feature Performance Summary</b>").hide(axis="index")
        
        # 异常熔断：如果筛选后为空，直接返回表框架，避免底图渲染报错
        if df.empty:
            return styler
        
        if "psi_max" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn_r", subset=["psi_max"], vmin=0, vmax=0.25)
            
        if "iv" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["iv"], vmin=0.02, vmax=0.2)
        if "auc" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["auc"], vmin=0.5, vmax=0.65)
        if "ks" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["ks"], vmin=5, vmax=20)

        if "rc_min" in df.columns:
            styler = styler.background_gradient(cmap="RdYlGn", subset=["rc_min"], vmin=0.5, vmax=1.0)
            
        if "mono" in df.columns:
            # coolwarm 色带: -1 为深蓝(单调递减)，0 为灰白(无单调性)，1 为深红(单调递增)
            styler = styler.background_gradient(cmap="coolwarm", subset=["mono"], vmin=-1, vmax=1)

        return styler.format("{:.4f}", subset=df.select_dtypes("number").columns)

    def show_trend(self, 
                   metric: str, 
                   features: Optional[Union[str, List[str]]] = None, # 新增特征筛选参数
                   group_ascending: bool = True, 
                   sort_by: Union[str, List[str]] = "Total", 
                   sort_ascending: bool = False) -> "pd.io.formats.style.Styler":
        """
        [Interactive] 展示指标的时间趋势热力图。

        渲染并返回一个带条件格式 (Conditional Formatting) 的 Pandas Styler 对象，
        用于直观分析特征在不同时间切片（或客群分组）下的指标波动趋势。内置了针对
        风控业务语义优化的专属色盘 (Colormap)。

        Parameters
       -------
        metric : str
            需要展示的指标名称。支持的选项可通过 `self.trend_tables.keys()` 查看
            (通常包含 'psi', 'auc', 'ks', 'iv', 'bad_rate', 'risk_corr')。
        features : str or List[str], optional
            需要展示的特征名列表。若为 None，则展示所有特征。
        group_ascending : bool, default True
            分组/时间切片列的排序方向 (横向)。True 表示正序（从左到右由旧到新 / 由小到大）。
        sort_by : str or List[str], default "Total"
            特征行的排序依据列。默认按照全局表现 (Total) 排序。
        sort_ascending : bool, default False
            特征行的排序方向 (纵向)。默认降序 (False)，即把表现最差/最好的特征排在最上面。

        Returns
       ----
        pd.io.formats.style.Styler
            渲染完成的热力图对象。在 Jupyter Notebook 环境下会自动渲染为精美表格。
        """
        if metric not in self.trend_tables:
            raise ValueError(f"Unknown metric: {metric}. Options: {list(self.trend_tables.keys())}")
        
        # 转换为 Pandas 副本进行安全的样式处理
        df: pd.DataFrame = self._to_pd(self.trend_tables[metric]).copy()
        
        # 特征筛选逻辑
        if features is not None:
            if isinstance(features, str):
                features = [features]
            df = df[df["feature"].isin(features)]

        # 行排序：紧跟 sort_by 和 sort_ascending 语义
        if sort_by in df.columns or (isinstance(sort_by, list) and all(c in df.columns for c in sort_by)):
            df = df.sort_values(by=sort_by, ascending=sort_ascending)
        
        # 识别列类型并重排时间切片列
        meta_cols = ["feature", "dtype"]
        special_cols = ["Total"]
        time_cols = [c for c in df.columns if c not in meta_cols + special_cols]
        
        # 列排序：受 group_ascending 控制
        time_cols_sorted = sorted(time_cols, reverse=not group_ascending)

        # 组装最终的列顺序：元数据 -> 时间切片 -> 汇总列
        final_cols = [c for c in meta_cols if c in df.columns] + \
                     time_cols_sorted + \
                     [c for c in special_cols if c in df.columns]
        df = df[final_cols]

        # 基础表格样式初始化
        styler = df.style.set_caption(f"<b>Trend Analysis: {metric.upper()}</b>").hide(axis="index")
        styler = styler.set_properties(subset=["feature"], **{'text-align': 'left', 'font-weight': 'bold'})

        if df.empty:
            return styler # 如果筛选后为空，直接返回空表格框架，避免报错

        # 根据不同业务指标的阈值与方向，映射专属渐变色盘
        if metric == "psi":
            styler = styler.background_gradient(
                cmap="RdYlGn_r", subset=time_cols_sorted, vmin=0, vmax=0.25, axis=None
            )
        elif metric in ["auc", "ks", "iv"]:
            styler = styler.background_gradient(
                cmap="RdYlGn", subset=time_cols_sorted, axis=None
            )
        elif metric == "bad_rate":
            styler = styler.background_gradient(
                cmap="Blues", subset=time_cols_sorted, axis=None
            )
        elif metric == "risk_corr":
            styler = styler.background_gradient(
                cmap="RdYlGn", subset=time_cols_sorted, vmin=0.5, vmax=1.0, axis=None
            )

        # 统一数值精度
        format_cols = [c for c in df.select_dtypes(include=[np.number]).columns]
        return styler.format("{:.4f}", subset=format_cols)


    def write_excel(self, path: str = "mars_bin_report.xlsx", engine: str = "openpyxl") -> None:
        """
        [Professional Export] 自动化导出分箱明细表。
        
        Parameters
       -------
        path : str
            导出的 Excel 文件路径。
        engine : str, default="openpyxl"
            写入 Excel 的底层引擎。
            - "auto": 自动检测，Win/Mac 下优先尝试 xlwings，若失败或在 Linux 下则回退至 openpyxl。
            - "xlwings": 强制使用 xlwings 引擎 (依赖本地安装的 Excel 应用程序，格式保留最完美)。
            - "openpyxl": 强制使用 openpyxl 引擎 (无需安装 Excel，跨平台兼容性好)。
        """
        valid_engines = ["auto", "xlwings", "openpyxl"]
        if engine not in valid_engines:
            raise ValueError(f"❌ 不支持的 engine: '{engine}'，请从 {valid_engines} 中选择。")

        # 智能定位模板路径
        package_name = "mars.analysis" 
        template_name_xlwings = "mars_bin_report_win_mac.xlsx"
        template_name_openpyxl = "mars_bin_report_linux.xlsx"
        
        def get_template_path(fname):
            try:
                import importlib.resources as resources
                with resources.as_file(resources.files(package_name).joinpath(fname)) as p:
                    return str(p)
            except Exception:
                return os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)

        # 配置内部参数
        START_WRITE_ROW = 4
        STYLE_SOURCE_ROW = 2
        FONT_NAME = "Microsoft YaHei"
        FONT_SIZE = 8
        SHEET_NAME = "分组明细"

        # 引擎解析与初始化
        use_xlwings = False

        if engine == "xlwings":
            use_xlwings = True
        elif engine == "openpyxl":
            use_xlwings = False
        else:  # "auto"
            # 自动探测环境
            is_gui_env = sys.platform.startswith("win") or sys.platform.startswith("darwin")
            use_xlwings = is_gui_env

        # 验证 xlwings 可用性
        if use_xlwings:
            try:
                import xlwings as xw
                # 测试 Excel 应用程序是否真正可用
                xw.App(visible=False, add_book=False).quit()
                template_path = get_template_path(template_name_xlwings)
            except Exception as e:
                if engine == "xlwings":
                    # 用户强制要求但失败，直接抛错
                    raise RuntimeError(f"❌ 强制使用 xlwings 引擎失败，请确认系统已正确安装 Excel 及 xlwings 库。\n报错详情: {e}")
                else:
                    # auto 模式下失败，降级处理
                    print(f"⚠️ xlwings 启动失败，将降级使用 openpyxl 引擎: {e}")
                    use_xlwings = False

        # 若未使用 xlwings，则准备 openpyxl 依赖
        if not use_xlwings:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            template_path = get_template_path(template_name_openpyxl)

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"❌ 找不到模板文件: {template_path}")

        # 准备数据
        df_pd = self._to_pd(self.detail_table)
        total_cols = len(df_pd.columns)

        # ================= 路径 A: xlwings 写入 =================
        if use_xlwings:
            app = None
            try:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False
                
                wb = app.books.open(template_path)
                ws = wb.sheets[SHEET_NAME]
                
                # 防止 Excel 将长数字字符串转为科学计数法
                if 'mars_group' in df_pd.columns:
                    df_pd['mars_group'] = "'" + df_pd['mars_group'].astype(str)
                
                # 写入数据
                ws.range((START_WRITE_ROW, 1)).value = df_pd.values
                final_row = START_WRITE_ROW + len(df_pd) - 1
                
                # 样式格式刷
                if final_row >= START_WRITE_ROW:
                    src_row = int(STYLE_SOURCE_ROW)
                    start_row = int(START_WRITE_ROW)
                    end_row = int(final_row)
                    max_col = int(total_cols)

                    source_range = ws.range((src_row, 1), (src_row, max_col))
                    data_range = ws.range((start_row, 1), (end_row, max_col))
                    
                    source_range.copy()
                    data_range.api.PasteSpecial(Paste=-4122) # xlPasteFormats
                
                # 统一字体
                full_range = ws.range((1, 1), (final_row, total_cols))
                full_range.api.Font.Name = FONT_NAME
                full_range.api.Font.Size = FONT_SIZE
                
                # 更新超级表 (ListObject)
                if ws.api.ListObjects.Count > 0:
                    table = ws.api.ListObjects(1)
                    new_ref = ws.range((1, 1), (final_row, total_cols)).get_address(False, False)
                    table.Resize(ws.range(new_ref).api)
                
                # 清理旧数据
                last_used_row = ws.api.UsedRange.Rows.Count
                if last_used_row > final_row:
                    ws.range(f"{final_row + 1}:{last_used_row}").api.Delete()

                wb.save(path)
                print(f"[xlwings Engine] 导出成功: {path}")

            except Exception as e:
                raise RuntimeError(f"xlwings 导出过程出错: {e}")
            finally:
                if 'wb' in locals() and wb: wb.close()
                if app: app.quit() 

        # ================= 路径 B: openpyxl 写入 =================
        else:
            wb = openpyxl.load_workbook(template_path)
            ws = wb[SHEET_NAME]
            
            mars_group_idx = -1
            if "mars_group" in df_pd.columns:
                mars_group_idx = list(df_pd.columns).index("mars_group") + 1
            
            # 提取并缓存样式模板
            style_map = {}
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=STYLE_SOURCE_ROW, column=c)
                style_map[c] = {
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format
                }

            # 写入数据
            rows = df_pd.values.tolist()
            for r_offset, row_data in enumerate(rows):
                current_row = START_WRITE_ROW + r_offset
                for c_offset, value in enumerate(row_data):
                    c_idx = c_offset + 1
                    cell = ws.cell(row=current_row, column=c_idx, value=value)
                    
                    # 应用样式
                    if c_idx in style_map:
                        s = style_map[c_idx]
                        cell.font = s["font"]
                        cell.border = s["border"]
                        cell.fill = s["fill"]
                        cell.alignment = s["alignment"]
                        cell.number_format = s["number_format"]
                    
                    # 日期列单独处理
                    if c_idx == mars_group_idx:
                        cell.number_format = "yyyy-mm-dd"

            final_row = START_WRITE_ROW + len(rows) - 1

            # 更新超级表范围并容错
            if hasattr(ws, 'tables') and ws.tables:
                new_ref = f"A1:{get_column_letter(total_cols)}{final_row}"
                for tbl_name in list(ws.tables.keys()):
                    tbl_obj = ws.tables[tbl_name]
                    
                    if hasattr(tbl_obj, 'ref'):
                        tbl_obj.ref = new_ref
                    else:
                        # 容错：处理 openpyxl 偶尔将 Table 解析为纯字符串的问题
                        del ws.tables[tbl_name]
                        new_tbl = Table(displayName=tbl_name, ref=new_ref)
                        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                        new_tbl.tableStyleInfo = style
                        ws.add_table(new_tbl)

            # 删除多余行
            if ws.max_row > final_row:
                ws.delete_rows(final_row + 1, ws.max_row - final_row)

            wb.save(path)
            print(f"[openpyxl Engine] 导出成功: {path}")